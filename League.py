import pandas as pd
import requests
import datetime 
import os
import numpy as np
from Player import *
import openpyxl as xl


"""
The League Class.
    This class describes the league in question
    Attributes:
        leagueID: the ID of this league
        years[]: a dictionary with keys of all the years this league has been active, and the value is the base URL for the espn API
        swid: the swid cookie of the user in this league            #used for logging in purposes
        espn_s2: the espn_s2 cookie of the user in this league      #used for logging in purposes
        dumpPath: the path that will be used to dump the cookies file
        cookieFile: a text file with the cookies associated with the fantasy account
        player_dict: A dictionary of players in this league
            Keys: Player name (string) Value: The Player object 
        player_array: an array of player names, which corresponds to the dictionary keys in the player_dict

"""
class League(object):
    def __init__(self, leagueID, yearlist=[],swid=None, espn_s2=None):
        self.leagueID=leagueID
        self.years= self.configYears(yearlist)
        self.SWID= swid
        self.espn_s2= espn_s2
        self.dumpPath= self.configPath() #Create a path to dump files

        try:
            self.cookieFile= open(str(self.leagueID)+"-Cookies.txt", "r")
            if self.cookieFile.readlines() == []:
                self.cookieFile.close()
                self.configCookies(str(self.leagueID)+"-Cookies.txt")
            else:
                self.cookieFile.close()
                self.cookieFile = open(str(self.leagueID)+"-Cookies.txt", "r")
                self.SWID = self.cookieFile.readline()
                self.SWID = self.SWID[:-1]
                self.espn_s2 = self.cookieFile.readline()
                self.cookieFile.close()
        except:
            self.configNewCookies()
        self.transArr=[]
        self.leagueName=''
        temp= self.configPlayers(2020)
        self.player_dict=temp[0]
        self.player_array=temp[1]
        self.linkTeamID()
        self.configExcelFile()
        


    #Configure a Cookies File that already exists, but has no/incorrect data in it  
    def configCookies(self, filename):
        self.cookieFile= open(filename, "w+")
        self.SWID=input("Provide your SWID key as a string (include {}):").rstrip()
        self.espn_s2=input("Provide your espn_s2 cookie as a string:").rstrip()
        self.cookieFile.write(self.SWID+"\n")
        self.cookieFile.write(self.espn_s2)
        self.cookieFile.close()

    #Configure a New Cookies File
    def configNewCookies(self):
        filename= os.path.join(self.dumpPath, str(self.leagueID)+"-Cookies.txt")
        self.configCookies(filename)   

    #Returns a path to save generated files
    def configPath(self): 
        return os.getcwd()

    #Returns a dictionary of years and the associated links with them
    def configYears(self, yearList):
        yeardict= {}
        for year in yearList:
            if year != 2020: 
                yeardict.update({year: "https://fantasy.espn.com/apis/v3/games/ffl/leagueHistory/"+str(self.leagueID)+"?seasonId="+str(year)})
            else:
                yeardict.update({year: "https://fantasy.espn.com/apis/v3/games/ffl/seasons/"+str(year)+"/segments/0/leagues/"+str(self.leagueID)})
        return yeardict


    #Returns a tuple containing a dictionary with all the player objects, and an array of all the dictionary keys
    def configPlayers(self, year):  #Make the keys all lowercase with one space between first and last time
        playerdict= {}
        playerArray=[]
        mTeam= self.getmTeam(year)
        teamArray= mTeam.get("members")
        for userDict in teamArray:
            name=(userDict.get('firstName')+' '+userDict.get('lastName'))
            name=self.makeName(name)
            playerArray.append(name) 
            playerdict.update({name: Player(userDict.get('firstName'), userDict.get('lastName'), userDict.get('id'), userDict.get('displayName'))})
            playerdict.update({userDict.get('id'): playerdict.get(name)})  #make the player obtainable by longID as well, not just name
        return (playerdict, playerArray)


    def configExcelFile(self):
        try: 
            self.openExcel()
        except:
            self.createExcel()



#===================================================================================================
# Various Getters for different ESPN view HTTP requests
#===================================================================================================

    #Returns the contents of the mStandings param page
    def getmStandings(self, year):
        url= self.years.get(year) + "?view=mStandings"
        if (self.SWID == '{}') and (self.espn_s2 == ''):
            r=requests.get(url).json()
        else:
            r= requests.get(url, cookies = {"swid": self.SWID,"espn_s2": self.espn_s2 }).json()
        return r

    def getmRoster(self, year):
        url= self.years.get(year) + "?view=mRoster"
        if (self.SWID == '{}') and (self.espn_s2 == ''):
            r=requests.get(url).json()
        else:
            r= requests.get(url, cookies = {"swid": self.SWID,"espn_s2": self.espn_s2 }).json()
        return r

    def getmBoxScore(self, year): 
        url= self.years.get(year) + "?view=mBoxScore"
        if (self.SWID == '{}') and (self.espn_s2 == ''):
            r=requests.get(url).json()
        else:
            r= requests.get(url, cookies = {"swid": self.SWID,"espn_s2": self.espn_s2 }).json()
        return r

    def getmTeam(self, year):
        url= self.years.get(year) + "?view=mTeam"
        if (self.SWID == '{}') and (self.espn_s2 == ''):
            r=requests.get(url).json()
        else:
            r= requests.get(url, cookies = {"swid": self.SWID,"espn_s2": self.espn_s2 }).json()
        return r 

    def getmSettings(self, year):
        url= self.years.get(year) + "?view=mSettings"
        if (self.SWID == '{}') and (self.espn_s2 == ''):
            r=requests.get(url).json()
        else:
            r= requests.get(url, cookies = {"swid": self.SWID,"espn_s2": self.espn_s2 }).json()
        return r

    def getmSchedule(self, year):
        url= self.years.get(year) + "?view=mSchedule"
        if (self.SWID == '{}') and (self.espn_s2 == ''):
            r=requests.get(url).json()
        else:
            r= requests.get(url, cookies = {"swid": self.SWID,"espn_s2": self.espn_s2 }).json()
        return r

    def getplayer_wl(self, year):
        url= self.years.get(year) + "?view=player_wl"
        if (self.SWID == '{}') and (self.espn_s2 == ''):
            r=requests.get(url).json()
        else:
            r= requests.get(url, cookies = {"swid": self.SWID,"espn_s2": self.espn_s2 }).json()
        return r


#=============================================================================================
# Helper Functions
#=============================================================================================

    #Converts the name of the team owner to a more standardized format
    def makeName(self, string):
        string=string.lower()
        i= string.find(' ')
        strings= string[:i+1] + string[i+1:].strip()
        return strings

    #adds teamID to each player and creates the teamID to phyID translation array
    def linkTeamID(self):
        mBoxScore= self.getmBoxScore(2020)
        self.leagueName= mBoxScore.get('settings').get('name').replace(" ", "")

        i=0
        #print('LeagueName: ' + self.leagueName)
        for team in mBoxScore.get('teams'):
            for owner in team.get('owners'):
                #print("This Owner: "+self.player_dict.get(owner).firstname+' '+self.player_dict.get(owner).lastname)
                self.player_dict.get(owner).setteamID(team.get('id'))
                self.player_dict.get(owner).setphyID(i)
            self.transArr.append( [team.get('id'), i] )
            i += 1
        
        
    def translate(self, teamID):
        for entry in self.transArr:
            if entry[0] == teamID:
                return entry[1]    
        print("No valid team!")
        
#===================================================================================================================
# Excel Configuration Functions
#===================================================================================================================

    #Create and save an excel workbook
    def createExcel(self):
        wb=xl.Workbook() 
        ws=wb.active
        ws.title= "Overview"
        wb.create_sheet("2020") #expand to all seasons later
        wb.save(os.getcwd()+"\\"+self.leagueName+ ".xlsx")

    def getWorkbook(self):
        return xl.load_workbook(os.getcwd()+"\\"+self.leagueName+".xlsx")
            
    def openExcel(self):
        wb=self.getWorkbook()
        wb.save(os.getcwd()+"\\"+self.leagueName+ ".xlsx")
    

#===================================================================================================================
# Season Excel Functions
#===================================================================================================================

    
# thisScores=[
#           [123, 123, 86, 104, 0, 0, 0, 0,...],  #this has phyID 0
#           [114, 146, 98, 123, 0, 0, 0, 0,...],  #this has phyID 1 ..
#           [                              ...],
#           [                              ...] ]
    def makeScoreArray(self,year):
        mStandings=self.getmStandings(year)
        status=mStandings.get('status')
        #make array of scores for the player 
        # print("teamsJoined: "+str(mStandings.get('teamsJoined')))
        # print("finalScoringPeriod: "+str(mStandings.get('finalScoringPeriod')))
        thisScores=np.zeros((status.get('teamsJoined'), status.get('finalScoringPeriod')))
        schedule= mStandings.get("schedule")
        for game in schedule:
            week= game.get('matchupPeriodId') - 1
            away= game.get('away')
            home= game.get('home')
            thisScores[self.translate(away.get('teamId'))][week]=away.get('totalPoints')
            thisScores[self.translate(home.get('teamId'))][week]=home.get('totalPoints')
        for playerName in self.player_array:
            player=self.player_dict.get(playerName)
            player.scores=thisScores[self.translate(player.teamID)]
        return thisScores

    def writeScoreArray(self, year):
        arr= self.makeScoreArray(year)
        wb=xl.load_workbook(os.getcwd()+"\\"+self.leagueName+".xlsx")
        ws=wb[str(year)]
        #Make column dimension correct
        ws.column_dimensions['A'].width= 20
        #Make Teams 
        for playerName in self.player_array:
            player= self.player_dict.get(playerName)
            ws.cell(self.translate(player.teamID) +2,1).value=playerName
        for i in range(len(arr)):
            sum=0
            count=0
            for j in range(len(arr[i])):
                if i==1:
                    ws.cell(1, j+2).value="Week "+ str(j+1)
                ws.cell(i+2,j+2).value=arr[i][j]
                sum += arr[i][j]
                if arr[i][j] != 0.0:
                    count+=1
            avg=sum/count
            ws.cell(i+2, len(arr[0])+3).value= avg
        ws.cell(1,len(arr[0])+3).value="Average"   
        wb.save(os.getcwd()+"\\"+self.leagueName+".xlsx")
        self.createLinechart(year, count, len(arr[0]))

    def createLinechart(self, year, numrows, numcols):
        wb=self.getWorkbook()
        ws=wb[str(year)]
        values = xl.chart.Reference(ws, min_col=2, min_row=2, max_col=2+numcols, max_row=2+numrows )
        frame=pd.read_excel(".\\"+self.leagueName+".xlsx", sheet_name= str(year))
        print(frame)
        thischart = xl.chart.LineChart()
        thischart.title="Scores vs. Week"
        thischart.y_axis.title= "Points"
        thischart.x_axis.title= "Weeks"
        thischart.add_data(frame.T)
        ws.add_chart(thischart,"B15")
        wb.save(os.getcwd()+"\\"+self.leagueName+".xlsx")

        



#===================================================================================================================
# Player Specific functions
#===================================================================================================================

    """
    What to run with the -p flag
        @param: player: string of the player to perform the action on"""
    def dashPscript(self, player):
        self.playerRecord(player)

    def playerRecord(self, player):
        #Get Worksheet
        wb=self.getWorkbook()
        ws= wb["Overview"]
        ws.cell(1,1).value= player 
        ws.cell(1,2).value= "Wins"
        ws.cell(1,3).value= "Losses"
        ws.cell(1,4).value= "Ties"
        #TODO Change Column width
        #Get Player
        mTeam=self.getmTeam(2020)
        pl= self.player_dict.get(player)
        owner= pl.longID
        teams=mTeam.get('teams')
        for team in teams:
            if owner in team.get("owners"):
                ovr= team.get("record").get("overall")
                pl.wins= ovr.get("wins")
                pl.losses= ovr.get("losses")
                pl.ties= ovr.get("ties")
                pl.ptsFor=ovr.get("pointsFor")
                pl.ptsAginst=ovr.get("pointsAgainst")
        ws.cell(2,1).value= "Record"
        ws.cell(2,2).value= str(pl.wins)
        ws.cell(2,3).value= str(pl.losses)  
        ws.cell(2,4).value= str(pl.ties)
        wb.save(os.getcwd()+"\\"+self.leagueName+".xlsx")

    




    




#Runs this code if this is a script
if __name__ == '__main__':
    myLeague= League(143434, [2018,2019,2020])
    myLeague.writeScoreArray(2020)
